from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
art=[]
name=[]
model=[]
price=[]
value=[]
href=[]

wb = Workbook()
ws = wb.active


html_temp=requests.get(r'https://spb.shinservice.ru/catalog/tyres/?filter=by-params&page=1&sort=stock')
soup_temp = BeautifulSoup(html_temp.text, 'html.parser')
pages_temp=[]
pages_temp.append(soup_temp.find(class_="summary__SummaryText-sc-10rltzp-0 cDBleI catalog-grid-summary"))
pages = pages_temp[0].text.split()
for page in range(int(pages[0])//36+1):

    html=requests.get(fr'https://spb.shinservice.ru/catalog/tyres/?filter=by-params&page={page}&sort=stock')

    soup = BeautifulSoup(html.text, 'html.parser')
    article = soup.find_all('article')

    for data in article:

        name.append(data.find(class_="title__TitleText-sc-1qgdk6a-2 jdAtzo stp-brand-model-title"))
        model.append(data.find(class_="goods-attribute-value"))
        price.append(data.find(class_="price-section__PriceDescription-sc-vpwglt-3 gcEGyG stp-price-total"))
        value.append((data.find(class_="availability-link__AvailabilityLinkText-sc-7as1wv-1 fvCduw")))
        art.append(data.find(class_="copy-sku-section__CopySkuSectionStyledSpan-sc-1dttp0z-1 eKKSHV catalog-card-sku-value"))
        href.append(data.find("a")["href"])
    print(len(href)//36 ,f"page out of {int(pages[0])//36+1}")

for i in range(len(art)):
    ws.cell(row=i+1, column=1).value=i
    ws.cell(row=i+1, column=2).value=art[i].text
    ws.cell(row=i+1, column=3).value = name[i].text
    ws.cell(row=i+1, column=4).value = model[i].text
    ws.cell(row=i+1, column=5).value = price[i].text
    ws.cell(row=i+1, column=6).value = value[i].text
    ws.cell(row=i+1, column=7).value = 'spb.shinservice.ru'+href[i]
wb.save("ShinService.xlsx")